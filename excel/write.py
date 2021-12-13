"""
    Write Excel File
"""
import openpyxl as xl
import traceback

import openpyxl.styles.fills as fills
from openpyxl.styles.fonts import Font
from openpyxl.styles import Color, PatternFill, Alignment, Border, Side

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))


def create_excel_file(data: list, sheets_info: dict, save_path: str):
    # workbook 생성
    wb = xl.Workbook()

    # first sheet(필수)
    sheet = wb.active
    sheet.title = "API Summary"

    # API 목록
    insert_summary(sheet, data)

    # extra sheets
    insert_sheets_info(wb, sheets_info)

    # set file path
    excel_file = save_path

    # save file
    wb.save(excel_file)
    wb.close()

    print("saved")


# API 목록 추가
def insert_summary(sheet, data: list):
    # set header
    headers = ['순번', '대메뉴', '중메뉴', '소메뉴', 'API URI', 'methods']
    for idx, name in enumerate(headers):
        header_cell = sheet.cell(1, idx + 1, name)
        set_header_style(header_cell, "center", True)

    current_row_number = 2
    current_col_number = 2

    # sheet data
    for r_num, row in enumerate(data):
        sheet.cell(current_row_number + r_num, current_col_number - 1, r_num + 1)
        for c_num, col in enumerate(row.split("|")):
            sheet.cell(current_row_number + r_num, current_col_number + c_num, col)

    # set sheet border
    set_border(sheet)


# 시트별 API 상세 내용 추가
def insert_sheets_info(wb: xl.Workbook, sheets_info: dict):
    for sheet_name in sheets_info.keys():

        _row = 0
        _col = 0

        # add sheets (대메뉴별 시트 생성)
        _sheet = wb.create_sheet(f'{sheet_name}({len(sheets_info[sheet_name])})')

        # idx 는 api 개수
        for info in sheets_info[sheet_name]:

            # add header cells
            info_names = ["API 명", "API 설명", "Method", "URI", "", "Security"]
            for idx, info_name in enumerate(info_names):
                header_cell = _sheet.cell(_row + idx + 1, 1, info_name)
                if info_name not in "":
                    set_header_style(header_cell, "left", True)

            _sheet.cell(_row + 1, 2, info.summary).border = border
            _sheet.cell(_row + 2, 2, info.description).border = border
            _sheet.cell(_row + 3, 2, info.method).border = border
            _sheet.cell(_row + 4, 2, info.uri).border = border
            _sheet.cell(_row + 6, 2, info.security).border = border

            # add request info
            if info.request is not None:
                body_cell = _sheet.cell(_sheet.max_row + 2, 1, "Request Body")
                body_cell.font = Font(bold=True)

                current_row = _sheet.max_row
                req_headers = ["Content-type", "Required"]
                for h_idx, info_name in enumerate(req_headers):
                    header_cell = _sheet.cell(current_row + h_idx + 1, 1, info_name)
                    set_header_style(header_cell, "left", False)

                _sheet.cell(body_cell.row + 1, 2, info.request.content_type).border = border
                _sheet.cell(body_cell.row + 2, 2, info.request.required).border = border

                # schema headers
                schema_headers = ["Name", "Title", "Required", "Type", "Default", "Description"]
                for h_idx, h_name in enumerate(schema_headers):
                    header_cell = _sheet.cell(body_cell.row + 4, h_idx + 1, h_name)
                    set_header_style(header_cell, "center", False)

                # schema detail
                insert_schema_value(_sheet, header_cell, info.request.schema)

            if info.parameters is not None:
                body_cell = _sheet.cell(_sheet.max_row + 2, 1, "URI Parameters")
                body_cell.font = Font(bold=True)

                # schema headers
                schema_headers = ["Name", "Title", "In", "Required", "Type", "Default", "Description"]
                for h_idx, h_name in enumerate(schema_headers):
                    header_cell = _sheet.cell(body_cell.row + 1, h_idx + 1, h_name)
                    set_header_style(header_cell, "center", False)

                # schema detail
                insert_param_schema_value(_sheet, header_cell, info.parameters)

            if info.responses is not None:
                body_cell = _sheet.cell(_sheet.max_row + 2, 1, "Responses")
                body_cell.font = Font(bold=True)
                current_row = _sheet.max_row
                res_row_num = body_cell.row
                for res_info in info.responses:
                    res_headers = ["Status Code", "Description", "Content-type"]
                    for h_idx, info_name in enumerate(res_headers):
                        header_cell = _sheet.cell(current_row + h_idx + 1, 1, info_name)
                        set_header_style(header_cell, "left", False)

                    _sheet.cell(res_row_num + 1, 2, res_info.status_code).border = border
                    _sheet.cell(res_row_num + 2, 2, res_info.description).border = border
                    _sheet.cell(res_row_num + 3, 2, res_info.content_type).border = border

                    # schema headers
                    schema_headers = ["Name", "Title", "Required", "Type", "Default", "Description"]
                    for h_idx, h_name in enumerate(schema_headers):
                        header_cell = _sheet.cell(res_row_num + 5, h_idx + 1, h_name)
                        set_header_style(header_cell, "center", False)

                    # schema detail
                    insert_schema_value(_sheet, header_cell, res_info.schema)

                    current_row = _sheet.max_row + 1
                    res_row_num = _sheet.max_row + 1

            _row = _sheet.max_row + 2


def insert_schema_value(_sheet, header_cell, schema):

    for s_idx, schema_info in enumerate(schema):
        detail_idx = header_cell.row + 1 + s_idx
        _sheet.cell(detail_idx, 1, schema_info.name).border = border
        _sheet.cell(detail_idx, 2, schema_info.title).border = border
        _sheet.cell(detail_idx, 3, schema_info.required).border = border
        _sheet.cell(detail_idx, 4, schema_info.type).border = border
        _sheet.cell(detail_idx, 5, None).border = border
        try:
            # 엑셀 파일 열기 시 수식 오류로 인해 임시 처리
            if schema_info.default is not None and "==" not in str(schema_info.default):
                _sheet.cell(detail_idx, 5, schema_info.default).border = border
        except ValueError as e:
            # print("insert_sheets_info Exception :", e)
            # traceback.print_exc()
            pass

        _sheet.cell(detail_idx, 6, schema_info.description).border = border


def insert_param_schema_value(_sheet, header_cell, schema):
    for s_idx, schema_info in enumerate(schema):
        detail_idx = header_cell.row + 1 + s_idx
        _sheet.cell(detail_idx, 1, schema_info.name).border = border
        _sheet.cell(detail_idx, 2, schema_info.title).border = border
        _sheet.cell(detail_idx, 3, schema_info.schema_in).border = border
        _sheet.cell(detail_idx, 4, schema_info.required).border = border
        _sheet.cell(detail_idx, 5, schema_info.type).border = border
        _sheet.cell(detail_idx, 6, None).border = border

        try:
            # 엑셀 파일 열기 시 수식 오류로 인해 임시 처리
            if schema_info.default is not None and "==" not in str(schema_info.default):
                _sheet.cell(detail_idx, 6, schema_info.default).border = border
            else:
                _sheet.cell(detail_idx, 6, None).border = border
        except ValueError as e:
            pass

        _sheet.cell(detail_idx, 7, schema_info.description).border = border


def set_header_style(header_cell, horizontal, bold):
    header_cell_color = "00C0C0C0"
    header_cell.font = Font(bold=bold)
    header_cell.alignment = Alignment(horizontal=horizontal, vertical="center")
    header_cell.fill = PatternFill(start_color=header_cell_color, end_color=header_cell_color,
                                   fill_type=fills.FILL_SOLID)
    header_cell.border = border


def set_border(ws):
    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border


if __name__ == '__main__':
    print("excel write start")
